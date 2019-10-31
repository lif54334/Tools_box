#!/usr/bin/env python

# -*- encoding: utf-8 -*-

'''
@Author  :   {lif54334}

@Software:   PyCharm

@File    :   doctorview.py

@Time    :   2018/10/8 12:46

@Desc    :

'''
import re

from openpyxl import load_workbook
from xlutils.copy import copy
import xlwt
import numpy as np
import xlrd


def sert():
    names = ["中药", "汤药", "针灸", "中草药", "中成药", "中医", "中药散", "煎服", "刮痧", "放血", "敷药"]
    data = xlrd.open_workbook('review.xlsx')  # 打开电影.xlsx文件读取数据
    table = data.sheets()[0]
    h_num = table.nrows
    num=[0]*h_num
    doctors=list()
    l_num = table.ncols
    # print(h_num,l_num)  # 输出表格行数列数
    for h in range(0, h_num):
        doctor=table.cell(h,0).value
        review = table.cell(h, 2).value
        # print(doctor,review)
        doctors.append(doctor)
        for name in names:
            if re.search(name, str(review)):
                # print(name)
                num[h]=1
                break
            else:
                pass
    print(len(doctors),len(num))
    with open("num.txt","w",encoding="utf-8")as f:
        for i in num:
            f.write(str(i)+"\n")
    # return doctors,num

def sort(arr1,arr2):
    all_np(arr1)





def all_np(arr):
    arr = np.array(arr)
    key = np.unique(arr)
    result = {}
    for k in key:
        mask = (arr == k)
        arr_new = arr[mask]
        v = arr_new.size
        result[k] = v
    return result




def main():
    sert()
    # sort(arr1,arr2)


if __name__ == '__main__':
    main()
